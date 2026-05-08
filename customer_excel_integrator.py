import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class CustomerExcelIntegrator:
    def __init__(self, config_file='config.json'):
        self.config = self._load_config(config_file)
        self.output_dir = self.config.get('output_dir', 'output')
        self.ensure_directory(self.output_dir)
    
    def _load_config(self, config_file):
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return self._get_default_config()
    
    def _get_default_config(self):
        return {
            "output_dir": "output",
            "categories": ["个人客户", "企业客户", "VIP客户"],
            "fields": {
                "基本信息": ["客户ID", "姓名", "性别", "年龄", "联系电话", "邮箱"],
                "联系信息": ["地址", "城市", "省份", "邮编"],
                "业务信息": ["客户类型", "注册日期", "消费金额", "购买产品"]
            }
        }
    
    def ensure_directory(self, path):
        if not os.path.exists(path):
            os.makedirs(path)
    
    def create_excel_structure(self, wb, category_name):
        ws = wb.create_sheet(title=category_name)
        
        headers = []
        for section, fields in self.config['fields'].items():
            headers.extend(fields)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        return ws
    
    def add_customer_data(self, ws, customer_data):
        next_row = ws.max_row + 1
        headers = []
        for section, fields in self.config['fields'].items():
            headers.extend(fields)
        
        for col, header in enumerate(headers, 1):
            value = customer_data.get(header, "")
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def process_customers(self, customers):
        category_files = {}
        
        for category in self.config['categories']:
            wb = Workbook()
            wb.remove(wb.active)
            self.create_excel_structure(wb, category)
            category_files[category] = wb
        
        for customer in customers:
            category = customer.get('客户类型', '其他')
            if category not in category_files:
                wb = Workbook()
                wb.remove(wb.active)
                self.create_excel_structure(wb, category)
                category_files[category] = wb
            ws = category_files[category].active
            self.add_customer_data(ws, customer)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_files = []
        
        for category, wb in category_files.items():
            filename = f"{category}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            output_files.append(filepath)
            print(f"已生成文件: {filepath}")
        
        return output_files
    
    def load_customers_from_json(self, json_file):
        with open(json_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def load_customers_from_csv(self, csv_file):
        import csv
        customers = []
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                customers.append(row)
        return customers
    
    def run(self, input_file):
        if input_file.endswith('.json'):
            customers = self.load_customers_from_json(input_file)
        elif input_file.endswith('.csv'):
            customers = self.load_customers_from_csv(input_file)
        else:
            raise ValueError("不支持的文件格式，请使用JSON或CSV文件")
        
        return self.process_customers(customers)

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='客户信息Excel整合工具')
    parser.add_argument('input', help='输入文件路径（JSON或CSV格式）')
    parser.add_argument('-c', '--config', default='config.json', help='配置文件路径')
    args = parser.parse_args()
    
    integrator = CustomerExcelIntegrator(args.config)
    
    try:
        output_files = integrator.run(args.input)
        print("\n处理完成！")
        for file in output_files:
            print(f"  - {file}")
    except Exception as e:
        print(f"处理失败: {e}")

if __name__ == '__main__':
    main()